from flask import (Flask,redirect,url_for,render_template,
                   request, flash, session, abort,
                   send_file, send_from_directory,
                   Response)

from webforms import (EndUserLicenseAgreement, 
                    LoginForm, Forget_pswForm, SignupForm,
                    Reset_pswForm, LogoutForm, SupplierForm,
                    BuyerForm, LineItemsForm, AddInfoForm)  

#### Flask Login ####
from flask_login import (LoginManager, login_user, logout_user,
                          current_user, login_required)
from werkzeug.security import (generate_password_hash, 
                                check_password_hash)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.sql import func
from flask_login import UserMixin

#### Flask mail ####
from datetime import datetime
import pytz
from flask_mail import Mail, Message

#### Google OAuth ####
from google_auth_oauthlib.flow import Flow
from pip._vendor import cachecontrol
from google.oauth2 import id_token
import google.auth.transport.requests

#### Standard libraries ####
import os
import pathlib
import requests
from pprint import pprint 
import re 

# import csv
from glob import glob

#### Flask upload ####
from werkzeug.utils import secure_filename

#### Handling document ####
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle

#### Handling database ####
from sqlalchemy import create_engine, text

app=Flask(__name__)

app.config['SECRET_KEY'] = "my super secret key that no one is supposed to know"

#### Flask Sqlalchemy ####
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///user.db?timeout=30'
# app.config["SQLALCHEMY_DATABASE_URI"] = "postgresql://user:password@host:port/database_name"
# app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL")

app.config['SQLALCHEMY_TRACK_MODIFICATION'] = False
app.config['SQLALCHEMY_ECHO'] = True

#### Flask Email ####
app.config['MAIL_SERVER']   = 'smtp.gmail.com'
app.config['MAIL_PORT']     = 465
app.config['MAIL_USERNAME'] = 'bhlohass@gmail.com'
app.config['MAIL_PASSWORD'] = 'ihmmpogbjigueoxs'
# app.config['MAIL_PASSWORD'] = 'Bhlohass@76'
app.config['MAIL_USE_TLS']  = False
app.config['MAIL_USE_SSL']  = True

# where we will store the uploaded files
app.config['UPLOADED_FORDER'] = 'static/upload'
app.config['MAX_CONTEXT _LENGTH'] = 16 * 1000 *1000

# set of allowed file extensions
ALLOWED_EXTENSIONS = {'xlsx'}

# Create the extension 
mail = Mail(app)
db = SQLAlchemy(app)

##### Create Model #####
class Users(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(20))
    email = db.Column(db.String(120), nullable=False, unique=True)
    password_hash = db.Column(db.String(256), nullable=True)
    auth_provider = db.Column(db.String(50), default='local')
    login_timestamp = db.Column(db.Text, nullable=True)
    logout_timestamp = db.Column(db.Text, nullable=True)
    user_agreement_timestamp = db.Column(db.Text, nullable=True)
    
    # New column for blacklist
    blacklisted = db.Column(db.Boolean, default=False, nullable=False)
    
def __init__(self, email, username, password_hash,
                login_timestamp=None, logout_timestamp=None,
                user_agreement_timestamp=None, blacklisted=False):
    self.email = email
    self.username = username
    self.password_hash = password_hash
    self.login_timestamp = login_timestamp
    self.logout_timestamp = logout_timestamp
    self.user_agreement_timestamp = user_agreement_timestamp
    self.blacklisted = blacklisted
  

# Create or update the database with the new column
with app.app_context():
    db.drop_all()
    db.create_all()
    
###### Flask login setting ######
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

###### Add Favicon ########
@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static', 'images'),
                'how.ico', mimetype='image/x-icon')

###### Login logic ########
@app.route("/", methods=['GET','POST'])
def index():
    form = LoginForm()
    return render_template("login.html", form=form)

@app.route('/login',methods=['GET','POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        # Query the user by email
        user = Users.query.filter_by(email = form.email.data).first()
        
        # Where user are exists
        if user:
            if user.password_hash:
                # Check if the password is correct
                if check_password_hash(user.password_hash, form.password.data):
                   
                     # Check if user is blacklisted
                    if user.blacklisted:
                        flash("😥 Sorry, you have violated the user terms and conditions. \
                              Your access is temporarily denied! \
                              Please contact the administrator.", 'danger')
                        return redirect(url_for('login'))  # Redirect back to login page if blacklisted
                
                   
                    # First-time login
                    if user.user_agreement_timestamp is None:  
                        # Log in the user
                        login_user(user)
                        flash("😄 Welcome! Please review the EULA.", 'success')
                        return redirect(url_for('eula'))
                    
                    # Set login timestamp
                    kuala_lumpur=pytz.timezone('Asia/Kuala_Lumpur')
                            
                    # Get current time in Kuala Lumpur
                    datetime_kuala_lumpur = datetime.now(kuala_lumpur)
                    
                    # Format the time in MM/DD/YYYY HH:MM:SS AM/PM
                    formatted_time = datetime_kuala_lumpur.strftime('%d/%m/%Y %I:%M:%S %p')
                    
                    # print("Login time (Kuala Lumpur):", formatted_time)
                
                    user.login_timestamp = formatted_time  
                    
                    # commit the change to the database
                    db.session.commit()
                    
                    # log the user in and redirect to the dashboard
                    login_user(user)
                    flash("😄 Yeah, you have logged in", 'success')
                    return redirect(url_for('dashboard'))
                    
                else:
                    flash("😥 Sorry, your password is incorrect!", 'danger')
                    # Clear the password field
                    form.password.data = ""
        # User does not exist        
        else:
            flash("😥 Sorry, user does not exist!", 'danger')
            # Clear both email and password fields
            form.email.data = ""
            form.password.data = ""
           
    return render_template('login.html', form=form)


@app.route('/eula', methods=['GET', 'POST'])
@login_required
def eula():
    form = EndUserLicenseAgreement()
    excel_path = 'static/eula/terms_conditions.xlsx'

    # Check if the EULA PDF exists
    if not os.path.exists(excel_path):
        flash("EULA file not found.", "danger")
        return render_template('eula.html', form=form, terms_text=None)
    
    try:
        # Read Excel file
        df = pd.read_excel(excel_path, header=None, dtype=str)
        terms_text = "\n".join(df[0].dropna().tolist())  # Join non-empty lines

        # Convert text wrapped with ** to bold using <strong>
        terms_text = re.sub(r'\*\*(.*?)\*\*', r'<strong>\1</strong>', terms_text)

    except Exception as e:
        flash(f"Error reading EULA: {str(e)}", "danger")
        terms_text = None
    
    if form.validate_on_submit():
        if 'submit' in request.form and request.form['submit'] == 'Submit':
            # Set the user agreement timestamp
            kuala_lumpur = pytz.timezone('Asia/Kuala_Lumpur')
            datetime_kuala_lumpur = datetime.now(kuala_lumpur)
            formatted_time = datetime_kuala_lumpur.strftime('%d/%m/%Y %I:%M:%S %p')

            current_user.user_agreement_timestamp = formatted_time
            current_user.login_timestamp = formatted_time  # Update login timestamp as well
            db.session.commit()
            
            flash("😄 Thank you for accepting the EULA. You are now in dashboard.", 'success')
            return redirect(url_for('dashboard'))
        elif 'cancel' in request.form and request.form['cancel'] == 'Cancel':
            # Log out the user on cancel
            # logout_user()
            flash("😄 You have successfully logout!", 'success')
            return redirect(url_for('login'))

    # Pass the PDF path to the template
    return render_template('eula.html', form=form, terms_text=terms_text)

@app.route('/dashboard', methods=['GET', 'POST'])
# Protect only the dashboard
@login_required
def dashboard():
    form = LogoutForm()
    print(f"Current user: {current_user}")
    is_admin = current_user.email == "admin@76" and \
        check_password_hash(current_user.password_hash, "admin@76")
    
    if form.validate_on_submit():
        # Handle logout logic here
        flash('😄 You have been logged out.', 'success')
        return redirect(url_for('login'))  # Redirect to login page after logout
    return render_template('dashboard.html', 
                           form=form,
                           is_admin=is_admin)

@login_manager.user_loader
def load_user(user_id):
    return Users.query.get(int(user_id))
       
# Create logout
@app.route('/logout', methods=['GET','POST'])
def logout():
    if current_user.is_authenticated:
        # Set logout timestamp
        kuala_lumpur=pytz.timezone('Asia/Kuala_Lumpur')
                
        # Get current time in Kuala Lumpur
        datetime_kuala_lumpur = datetime.now(kuala_lumpur)
        
        # Format the time in MM/DD/YYYY HH:MM:SS AM/PM
        formatted_time = datetime_kuala_lumpur.strftime('%d/%m/%Y %I:%M:%S %p')
        
        # print("Logout time (Kuala Lumpur):", formatted_time)
        
        # Store the formatted string in the database
        current_user.logout_timestamp = formatted_time
        
        # Then commit the logout timestamp to the database
        db.session.commit()
        
        # Clear form data on logout
        session.pop('form_data', None)  
        session.pop('l_inv_no', None)
        session.pop('l_date_time', None)
        
        ########## Delete Excel Sheet ##########
        file_path = 'static/batch_submission/Multiple_transaction.xlsx'
        workbook = load_workbook(filename=file_path)
        print(f"'{file_path}' loaded successfully.")

        # Iterate backward from row 100 to row 6
        for row_num in range(100, 5, -1):  # Start at 100, stop before 5, decrement by 1
            doc_sheet = workbook["Documents"]
            doc_sheet.delete_rows(row_num)
            doc_line_sheet = workbook["DocumentLineItems"]
            doc_line_sheet.delete_rows(row_num)
            line_items_class_sheet = workbook["LineItemsAddClassifications"]
            line_items_class_sheet.delete_rows(row_num)
            line_items_tax_sheet = workbook["LineItemsTaxes"]
            line_items_tax_sheet.delete_rows(row_num)
            doc_total_tax_sheet = workbook["DocumentTotalTax"]
            doc_total_tax_sheet.delete_rows(row_num)
            line_item_dis_sheet = workbook["LineItemsDiscounts"]
            line_item_dis_sheet.delete_rows(row_num)
            doc_dis_sheet = workbook["DocumentDiscounts"]
            doc_dis_sheet.delete_rows(row_num)
            line_item_fee_sheet = workbook["LineItemsCharges"]
            line_item_fee_sheet.delete_rows(row_num)
            doc_fee_sheet = workbook["DocumentCharges"]
            doc_fee_sheet.delete_rows(row_num)

        # Save changes to the workbook
        workbook.save(filename=file_path)
        # print("Rows 6 to 100 have been deleted.")
        ########################################
        
        # Delete all files in 'static/upload/' folder
        upload_folder = 'static/upload/'
        for filename in os.listdir(upload_folder):
            file_path = os.path.join(upload_folder, filename)
            try:
                if os.path.isfile(file_path):
                    os.remove(file_path)
            except Exception as e:
                # print(f"Error deleting file {file_path}: {e}")
                flash('Error deleting file {file_path}: {e}')

    logout_user()
    session.clear()
    flash("😄 Thank you, you have logout and all data deleted completely!", 'success')
    return redirect(url_for('index'))

###### Signup logic ######
@app.route('/sign_up', methods=['GET', 'POST'])
def sign_up():
    form = SignupForm()
    if request.method == 'POST':
        # when field not fill out
        if not request.form['username'] or \
           not request.form['email'] or \
           not request.form['password'] or \
           not request.form['confirm_password']:
               flash("😥 Please fill out all the fields", 'danger')
        
        else:
            # however when field have been fill out 
            username = form.username.data
            email = form.email.data
            password = form.password.data

            # check if email already exists
            existing_user = Users.query.filter_by(email=email).first()
            if existing_user:
                flash("😥 This email is already registered", 'danger')
                return render_template('sign_up.html', form=form)
            
            # hash password
            hashed_pw = generate_password_hash(password)
            # Create user while ignoring login_timestamp, logout_timestamp, and user_agreement_timestamp
            user = Users(
                username=username,
                email=email,
                password_hash=hashed_pw
                # login_timestamp, logout_timestamp, user_agreement_timestamp are omitted
            )
           
            #### Database add user ####
            db.session.add(user)
            db.session.commit()
            flash('😄 User was added!', 'success')
            return redirect(url_for('login'))
    
    # if not success remain in new page
    return render_template('sign_up.html', form = form)    

@app.route('/forget', methods=['GET', 'POST'])
def forget():
    form = Forget_pswForm()
    # This method checks for form validation
    if form.validate_on_submit():
        # Retrieve username and email from the form
        username = form.username.data
        email = form.email.data
        
        # Store the username in session for the reset process
        # Store the username securely
        session['username'] = username
        session['email'] = email
        
        return redirect(url_for('send_email', email=email, 
                                username = username)) 
    # If the method is GET or the form is not valid, render the form    
    return render_template('forget_psw.html', form=form)

@app.route('/send_email/<email>', methods=['GET', 'POST'])
def send_email(email):
    msg_title = "Reset Password"
    sender = app.config['MAIL_USERNAME']  # Sender's email
    recipients = [email]  # The intended recipient email passed in the URL

    # Generate the reset link for the recipient
    username = session.get('username')  # Retrieve username from session
    reset_url = url_for('reset_psw', username=username, _external=True)
    print(f"Reset URL: {reset_url}")  # Debug statement
    print(f"Session data: {session}")  # Debug statement
    # Create the email body
    msg_body = f"Thank you for using How E-invoice.\nPlease click the link below to reset your password:\n{reset_url}"
    data = {
        'owner': "How App & Web",
        'title': msg_title,
        'body': msg_body,
    }

    # Configure the email message
    msg = Message(msg_title, sender=sender, recipients=recipients)
    msg.html = render_template("email.html", data=data)

    try:
        mail.send(msg)
        flash("😄 Reset Password Emailed", "success")
    except Exception as e:
        print(e)
        flash("😥 The email could not be sent. Please try again later.", "danger")

    return redirect(url_for('login'))

@app.route('/reset_psw', methods=['GET', 'POST'])
def reset_psw():
    form = Reset_pswForm()
    
    # Retrieve the username from the query parameters
    username = request.args.get('username')
    print(f"Username from query parameters: {username}")  # Debug statement
    
    if not username:
        print("Username is None or invalid. Redirecting to forget.")  # Debug statement
        flash('😥 Invalid or expired reset request', 'danger')
        return redirect(url_for('forget'))
    
    # Check if the user exists (case-insensitive and trimmed)
    user = Users.query.filter(Users.username.ilike(username.strip())).first()
    # print(f"User found in database: {user}")  # Debug statement
    
    if not user:
        # print("User not found in database. Redirecting to forget.")  # Debug statement
        flash('😥 User not found', 'danger')
        return redirect(url_for('forget'))

    if form.validate_on_submit():
        password = form.password.data
        confirm_password = form.confirm_password.data
        
        if password != confirm_password:
            flash("😥 Passwords do not match!", 'danger')
            return render_template('reset_psw.html', form=form)

        # Hash the new password and update the user's password
        hashed_pw = generate_password_hash(password)
        user.password_hash = hashed_pw
        db.session.commit()
        
        flash('😄 Password has been successfully updated!', 'success')
        return redirect(url_for('login'))
    
    # If the method is GET or form is invalid, render the reset form
    print("Rendering reset_psw.html template.")  # Debug statement
    return render_template('reset_psw.html', form=form)
######## Login with google ########
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID")
client_secrets_file = os.path.join(pathlib.Path(__file__).parent, "client_secret.json")

os.environ["OAUTHLIB_INSECURE_TRANSPORT"] = "1"
flow = Flow.from_client_secrets_file(
    client_secrets_file=client_secrets_file,
    scopes=["https://www.googleapis.com/auth/userinfo.profile", 
            "https://www.googleapis.com/auth/userinfo.email", 
            "openid"],
    redirect_uri="https://how-e-invoice.onrender.com/callback",
)


def login_is_required(function):
    def wrapper(*args, **kwargs):
        if "google_id" not in session:
            # Authorization required
            return abort(401)  
        else:
            return function()
    return wrapper

@app.route('/google_login')
def google_login():
    authorization_url, state = flow.authorization_url()
    session['state'] = state
    # print("Click Google")
    return redirect(authorization_url)

@app.route('/callback')
def callback():
    flow.fetch_token(authorization_response=request.url)

    if 'state' not in session or session['state'] != request.args.get('state'):
        return "Error: Invalid state parameter", 400

    credentials = flow.credentials
    # Fetch user info from Google
    user_info = get_google_user_info(credentials)
    
    # Extract and print the user's name and email
    user_name = user_info.get('name', 'Unknown User')
    user_email = user_info.get('email', 'No Email')

    # print(f"Google User Name: {user_name}")
    # print(f"Google User Email: {user_email}")
    
    #### For the timestamp ####
    kuala_lumpur=pytz.timezone('Asia/Kuala_Lumpur')
    # Get current time in Kuala Lumpur
    datetime_kuala_lumpur = datetime.now(kuala_lumpur)
    
    # Format the time in MM/DD/YYYY HH:MM:SS AM/PM
    formatted_time = datetime_kuala_lumpur.strftime('%d/%m/%Y %I:%M:%S %p')
    # print("Login time (Kuala Lumpur):", formatted_time)

    # Check if the user already exists in the database by email
    user = Users.query.filter_by(email=user_email).first()
    if not user:
        # If the user does not exist, create a new user
        user = Users(
            email=user_email,
            username=user_name,  
            # Leave password_hash as None
            password_hash=None,
            auth_provider='google'
        )
        db.session.add(user)
    # else:
    #     # Update login timestamp for existing user
    #     user.login_timestamp = formatted_time
        
    # Commit the changes
    db.session.commit()

    # Log the user in
    login_user(user)
    
    # Check if user is blacklisted
    if user.blacklisted:
        flash("😥 Sorry, you have violated the user agreement. \
                Your access is temporarily denied! \
                Please contact the administrator.", 'danger')
        return redirect(url_for('login'))  # Redirect back to login page if blacklisted
    
    # Redirect based on login status
    if user.user_agreement_timestamp is None:
        # First-time login: Redirect to EULA page
        flash("😄 Welcome! Please review the EULA.", 'success')
        return redirect(url_for('eula'))
    else:
        # Subsequent logins: Redirect to dashboard
        flash("😄 Logged in with Google!", 'success')
    return redirect(url_for('dashboard'))
      
def get_google_user_info(credentials):
    """Get user information from Google using OAuth credentials."""
    # Define the Google OAuth2 API URL for user info
    user_info_url = 'https://www.googleapis.com/oauth2/v2/userinfo'

    # Prepare the request headers with the access token
    headers = {
        'Authorization': f'Bearer {credentials.token}'
    }

    # Send a request to the Google API to get user info
    response = requests.get(user_info_url, headers=headers)

    # If the request is successful, return the user info as JSON
    if response.status_code == 200:
        return response.json()
    else:
        raise Exception(f"Failed to retrieve user info: {response.status_code}, \
                        {response.text}")
    
######## Excel upload ########
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() \
           in ALLOWED_EXTENSIONS
        
@app.route('/upload', methods=['GET', 'POST'])
def upload():
    form = LogoutForm()
    data = []
    if request.method == 'POST':
        # Check if the post request has the file part
        if 'filename' not in request.files:
            flash('😥 No file part', 'danger')
            return redirect(request.url)

        file = request.files['filename']
        
        if file.filename == '':
            flash('😥 No excel file selected for upload', 'danger')
            return redirect(url_for('dashboard'))
           
        elif file and allowed_file(file.filename):
            upload_folder = app.config['UPLOADED_FORDER']
            # f = request.files['filename']
            filepath = os.path.join(app.config['UPLOADED_FORDER'], file.filename)
            
            # Check if any Excel files already exist in the folder
            for existing_file in os.listdir(upload_folder):
                if existing_file.endswith('.xlsx'):
                    os.remove(os.path.join(upload_folder, existing_file))
            
            # Save the new file
            file.save(filepath)
            # print(file)
            flash('😄 Excel file successfully uploaded', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('😥 Only accept an excel file!', 'danger')
            return redirect(url_for('dashboard'))
            
    return render_template("dashboard.html", form=form)

@app.route('/download_1')
def download_1():
    path = "static/download/individual_e-invoice.xlsx"
    return send_file(path, as_attachment=True)

@app.route('/download_2')
def download_2():
    path = "static/download/consolidated_e-invoice.xlsx"
    return send_file(path, as_attachment=True)

###### Error handler ######
@app.errorhandler(404)
def invalid_route(e):
   return render_template('404.html') 

###### Supplier Page #####
@app.route('/supplier')
@login_required
def supplier():
    form = SupplierForm()
    
    excel_files = glob('static/upload/*.xlsx')
    
    # Check if any excel file exists
    if not excel_files:
        flash("😥 No excel file uploaded", "danger")
        return render_template('dashboard.html', form=form)

    # Use the first Excel file found in the folder
    file_path = excel_files[0]

    # Check the file type
    if not file_path.endswith('.xlsx'):
        flash("😥 Only Excel files are allowed", "danger")
        return render_template('dashboard.html', form=form)
    
    try:
        # Try to load the Excel file
        df = pd.read_excel(file_path, sheet_name=0)
        # print(df.iloc[:,1].to_string(index=False))
        
        # Replace NaN values with "NA"
        df = df.where(pd.notna(df), "NA")
        
        # Extract row data, mapping them to specific fields
        data = {
            # Adjust column index as needed
            "s_name": df.iloc[0, 1], 
            "s_id_type": df.iloc[1, 1],
            "s_business_no": df.iloc[3, 1],
            "s_msic_code": df.iloc[4, 1],
            "s_telephone": df.iloc[6, 1],
            "s_email": df.iloc[8, 1],
            
            "s_tin" : df.iloc[2, 1],
            "s_msic_desc":df.iloc[5, 1],
            "s_sst": df.iloc[7, 1] if pd.notna(df.iloc[6, 1]) else "NA",
            "s_tt": df.iloc[9, 1] if pd.notna(df.iloc[8, 1]) else "NA",
            "s_add_1": df.iloc[10, 1],
            "s_add_2": df.iloc[11, 1],
            "s_add_3": df.iloc[12, 1],
            "s_postcode": df.iloc[13, 1],
            's_city' : df.iloc[14, 1],
            "s_state": df.iloc[15, 1],
            "s_country": df.iloc[16, 1],
        }
        print(data)
        # Set each form field's data
        form.s_name.data = data["s_name"]
        form.s_business_no.data = data["s_business_no"]
        form.s_msic_code.data = data["s_msic_code"]
        form.s_telephone.data = data["s_telephone"]
        form.s_email.data = data["s_email"]
        
        form.s_tin.data = data["s_tin"]
        form.s_msic_desc.data = data["s_msic_desc"]
        form.s_sst.data = data["s_sst"]
        form.s_tt.data = data["s_tt"]
        form.s_add_1.data = data["s_add_1"]
        form.s_add_2.data = data["s_add_2"]
        form.s_add_3.data = data["s_add_3"]
        form.s_city.data = data["s_city"]
        form.s_postcode.data = data["s_postcode"]
        form.s_state.data = data["s_state"]
        form.s_country.data = data["s_country"]
        
        # store in session
        session['s_tin'] = data["s_tin"]
        session['s_name'] = data["s_name"]
        session['s_id_type'] = data["s_id_type"]
        session['s_business_no'] = data["s_business_no"]
        session['s_telephone'] = data["s_telephone"]
        session['s_email'] = data["s_email"]
        session['s_msic_code'] = data["s_msic_code"]
        session['s_msic_desc'] = data["s_msic_desc"]
        session['s_add_1'] = data["s_add_1"]
        session['s_add_2'] = data["s_add_2"]
        session['s_add_3'] = data["s_add_3"]
        session['s_city'] = data["s_city"]
        session['s_postcode'] = data["s_postcode"]
        session['s_state'] = data["s_state"]
        session['s_country'] = data["s_country"]
        
        # Disable all fields after populating them
        form.s_name.render_kw = {'disabled': True}
        form.s_business_no.render_kw = {'disabled': True}
        form.s_msic_code.render_kw = {'disabled': True}
        form.s_telephone.render_kw = {'disabled': True}
        form.s_email.render_kw = {'disabled': True}
        form.s_tin.render_kw = {'disabled': True}
        form.s_msic_code.render_kw = {'disabled': True}
        form.s_msic_desc.render_kw = {'disabled': True}
        form.s_sst.render_kw = {'disabled': True}
        form.s_tt.render_kw = {'disabled': True}
        form.s_add_1.render_kw = {'disabled': True}
        form.s_add_2.render_kw = {'disabled': True}
        form.s_add_3.render_kw = {'disabled': True}
        form.s_city.render_kw = {'disabled': True}
        form.s_postcode.render_kw = {'disabled': True}
        form.s_state.render_kw = {'disabled': True}
        form.s_country.render_kw = {'disabled': True}
        
        flash('😄 All fields was populated', 'success')
        return render_template('supplier.html', form=form)  # Success path

    except Exception as e:
        flash("😥 An error occurred while loading the Excel file", "danger")
        # print(f"Error loading Excel file: {e}")
        return render_template('supplier.html', form=form)
    

@app.route('/download_excel')
def download_excel():
    # Retrieve supplier name and date from session
    supplier_name = session.get('s_name', 'Supplier')  # Default to 'Supplier' if not set
    date_str = datetime.now().strftime('%Y-%m-%d')     # Format current date as YYYY-MM-DD

    # Create a dynamic file name
    file_name = f"{supplier_name}_{date_str}.xlsx".replace(' ', '_')  # Replace spaces with underscores

    # Path to the original file
    file_commit = 'static/batch_submission/Multiple_transaction.xlsx'

    # Send file with dynamic name
    session['fields_disabled'] = False
    
    return send_file(file_commit,
                    as_attachment=True,
                    download_name=file_name)

@app.route('/start_new_buyer')
@login_required
def start_new_buyer():
    # Clear buyer-related session data
    session.pop('form_data', None)
    session.pop('line_items', None)
   

    # Retrieve the last invoice number or set a default
    l_inv_no = session.get('l_inv_no', None)

    print(f"Before increment - last_inv_no: {l_inv_no}")  # Debugging print

    # Extract prefix and numeric part
    prefix = ''.join([c for c in l_inv_no if not c.isdigit()])
    numeric_part = ''.join([c for c in l_inv_no if c.isdigit()])
    next_number = int(numeric_part) + 1 if numeric_part.isdigit() else 1
    new_inv_no = f"{prefix}{next_number}"
    print('New Number:,{new_inv_no}')

    # Update session with new invoice number
    session['l_inv_no'] = new_inv_no

    print(f"After increment - new_inv_no: {session.get('l_inv_no')}")  # Debugging print

    # Enable fields for new input
    session['fields_disabled'] = False

    return redirect(url_for('buyer'))

@app.route('/buyer')
@login_required
def buyer():
    form = BuyerForm()
    
    # Get the latest invoice number from the session
    l_inv_no = session.get('l_inv_no', '')

    print(f"Buyer Page - l_inv_no: {l_inv_no}")  # Debugging print
    print(f"Session Data: {session}")

    # Pre-populate form fields if available
    if 'form_data' in session:
        form_data = session['form_data']
        form.b_id_type.data = form_data.get('b_id_type')
        form.b_id_no.data = form_data.get('b_id_no')
        form.b_tel_no.data = form_data.get('b_tel_no')
        form.b_email.data = form_data.get('b_email')
        form.b_name.data = form_data.get('b_name')
        form.b_add_1.data = form_data.get('b_add_1')
        form.b_add_2.data = form_data.get('b_add_2')
        form.b_add_3.data = form_data.get('b_add_3')
        form.b_city.data = form_data.get('b_city')
        form.b_country.data = form_data.get('b_country')
        form.b_tin.data = form_data.get('b_tin')
        form.b_sst.data = form_data.get('b_sst')
        form.b_tt.data = form_data.get('b_tt')
        form.b_postcode.data = form_data.get('b_postcode')
        form.b_state.data = form_data.get('b_state')
        form.b_search.data = form_data.get('b_search')
        form.b_result.data = form_data.get('b_result')

    return render_template("buyer.html", form=form, 
                           l_inv_no=l_inv_no)

@app.route('/clear', methods=['POST', 'GET'])
@login_required
def clear():
    form = BuyerForm()
    form.b_id_type.data = ''
    form.b_id_no.data = ''
    form.b_tel_no.data = ''
    form.b_email.data = ''
    form.b_add_1.data = ''
    form.b_add_2.data = ''
    form.b_add_3.data = ''
    form.b_city.data = ''
    form.b_country.data = ''
    form.b_tin.data = ''
    form.b_name.data = ''
    form.b_sst.data = ''
    form.b_tt.data = ''
    form.b_postcode.data = ''
    form.b_state.data = ''
    # print("Clicked clear button")
    
    # Clear form data on logout
    session.pop('form_data', None)  
    # Render with cleared form data instead of redirecting
    return render_template("buyer.html", 
                           form=form)

@app.route('/search', methods=['POST', 'GET'])
@login_required
def search():
    form = BuyerForm(is_search=True)
    # Default to 'buyer' if not specified
    page_type = request.args.get('page_type', 'New Buyer')
    
    # Initialize the variable at the beginning
    search_successful = False
    
    if form.validate_on_submit():
        # Extract the search term from the form field
        buyer_search = form.b_search.data 
        # print(f"Searching for: {buyer_search}")  # Debugging
        
        buyer_result = form.b_result.data
        # print(f"Searching for: {buyer_result}")
        
        if not buyer_search:
            flash("😥 Please fill in the search input", "danger")
            return redirect(url_for('buyer'))
        
        # print(f"Searching for: {buyer_search}")  # Debugging
        
        # Load the Excel file
        excel_files = glob('static/upload/*.xlsx')
        
        if not excel_files:
            flash("😥 No Excel files found", "danger")
            return redirect(url_for('buyer'))
    
        # Use the first Excel file found in the folder
        file_path = excel_files[0]
        
        try:
            # Try to load the Excel file
            df = pd.read_excel(file_path, sheet_name=1, dtype={
                'Telephone Number*': str,
                'Tax Identification Number (TIN)*': str,
                'Registration Number*': str,
                'Postcode*': str,
                # Add more fields if needed
            })
            # Filter the DataFrame by the search term in either "Alias" or "Name*" columns
            result = df[df['Alias'].str.contains(buyer_search, case=False, na=False) |
                        df['Name*'].str.contains(buyer_search, case=False, na=False)]
            # print(result)
            
            # Replace NaN values with "NA"
            df = df.where(pd.notna(df), "NA")
            
             # Check for duplicate names in the results
            if len(result) > 1:
                flash("😥 Duplicate name found. \
                       Input the correct alias or buyer name", "danger")
                return redirect(url_for('buyer'))
            
            # Check if there are any results and print the "Identification Card No."
            if not result.empty:
                # Accessing the first match
                id_type = result['ID Type* (Drop box)'].iloc[0]  
                print(f"Identification Card No.: {id_type}")
                
                tin_no = result['Tax Identification Number (TIN)*'].iloc[0] 
                # print(f"Tin No.: {tin_no}")
                
                reg_no = result['Registration Number*'].iloc[0]
                if pd.isna(reg_no):
                    reg_no = 'NA'
                # print(f"Registration/Identification/Passport Number*: {reg_no}")
                
                name = result['Name*'].iloc[0] 
                # print(f"Name: {name}")
                
                tel = result['Telephone Number*'].iloc[0]
                if pd.isna(tel):
                    tel = "NA" 
                print(f"Telephone number: {tel}")
                
                sst_no = result['SST Registration Number*'].iloc[0] 
                # Check if the value is NaN and replace it with "NA"
                if pd.isna(sst_no):
                    sst_no = "NA"
                # print(f"SST No.: {sst_no}")
                
                email = result['E-mail'].iloc[0] 
                # print(f"E-mail: {email}")
                
                tt_no = result['Tourism Tax Registration Number*'].iloc[0]  
                # Check if the value is NaN and replace it with "NA"
                if pd.isna(tt_no):
                    tt_no = "NA"
                # print(f"Tourism Tax: {tt_no}")
                
                add_1 = result['Address Line 1*'].iloc[0] 
                # print(f"Address 1: {add_1}")
                
                add_2 = result['Address Line 2'].iloc[0]  
                # print(f"Address 2: {add_2}")
                
                add_3 = result['Address Line 3'].iloc[0] 
                 # Check if the value is NaN and replace it with "NA"
                if pd.isna(add_3):
                    add_3 = "NA"
                # print(f"Address 3: {add_3}")
                
                city = result['City*'].iloc[0] 
                # print(f"City: {city}")
                
                postcode = result['Postcode*'].iloc[0]
                postcode = str(postcode) if pd.notna(postcode) else None 
                # print(f"Postcode: {postcode}")
                
                country = result['Country* (Drop box)'].iloc[0] 
                # print(f"Country: {country}")
                
                state = result['State* (Drop box)'].iloc[0]  
                # print(f"State: {state}")
                
                # Set each form field's data
                form.b_id_type.data = id_type
                form.b_id_no.data = reg_no
                form.b_tel_no.data = tel
                form.b_email.data = email
                form.b_add_1.data = add_1
                form.b_add_2.data = add_2
                form.b_add_3.data = add_3
                form.b_city.data = city
                form.b_country.data = country
                form.b_tin.data = tin_no
                form.b_name.data = name
                form.b_sst.data = sst_no
                form.b_tt.data = tt_no
                form.b_postcode.data = postcode
                form.b_state.data = state
                form.b_result.data = buyer_result  # Set form data
                
                # Disable all fields after populating them
                form.b_id_type.render_kw = {'disabled': True}
                form.b_id_no.render_kw = {'disabled': True}
                form.b_tel_no.render_kw = {'disabled': True}
                form.b_email.render_kw = {'disabled': True}
                form.b_add_1.render_kw = {'disabled': True}
                form.b_add_2.render_kw = {'disabled': True}
                form.b_add_3.render_kw = {'disabled': True}
                form.b_city.render_kw = {'disabled': True}
                form.b_country.render_kw = {'disabled': True}
                form.b_tin.render_kw = {'disabled': True}
                form.b_name.render_kw = {'disabled': True}
                form.b_sst.render_kw = {'disabled': True}
                form.b_tt.render_kw = {'disabled': True}
                form.b_postcode.render_kw = {'disabled': True}
                form.b_state.render_kw = {'disabled': True}
                
                # # Store data in session
                session['form_data'] = {
                    'b_id_type': id_type,
                    'b_id_no': reg_no,
                    'b_tel_no': tel,
                    'b_email': email,
                    'b_add_1': add_1,
                    'b_add_2': add_2,
                    'b_add_3': add_3,
                    'b_city': city,
                    'b_country': country,
                    'b_tin':  tin_no,
                    'b_name': name,
                    'b_sst': sst_no,
                    'b_tt': tt_no,
                    'b_postcode': postcode,
                    'b_state': state,
                    'b_search': buyer_search,  # store search term
                }
                
                # Extract the name for the supplier and buyer display
                df = pd.read_excel(file_path, sheet_name=0)
                s_name = df.iloc[0, 1]
                
                # Set `b_result` with both names
                form.b_result.data = f"Supplier name:  {s_name}\nBuyer name:  {name}"
                
                # Store `search_successful` in session
                session['search_successful'] = True 
                
                # Set the success flag to True to enable the Next button
                search_successful = True
                # print(f"Search successful, result: {form.b_result.data}")  # Debugging
                
                flash("😄 Alright your search found. Click 'Next' button to proceed!", "success")
            else:         
           
                form.b_result.data = "No matches found."
                # Clear session if no results found
                session.pop('search_successful', None)  
                # Reset the flag
                search_successful = False  
        except Exception as e:
            flash(f"😥 Error reading Excel file: {e}", "danger")
            
    return render_template('buyer.html', 
                           form=form, 
                           search_successful=search_successful)
    
@app.route('/add_line', methods=['POST', 'GET'])
@login_required
def add_line():
    # Initialize the form
    form = LineItemsForm()
    
    # Get existing line items from session or initialize a new list
    line_items = session.get('line_items', [])
    line_class_data = session.get('line_class_data', {
        'l_class_code': '',
        'l_desc': '',
        'l_measurement': '',
        'l_unit_price': ''
    })
    
    # Prepopulate form fields from session data
    # print("Prepopulating form with session data:", line_class_data)
    form.l_class_code.data = line_class_data.get('l_class_code', '')
    form.l_desc.data = line_class_data.get('l_desc', '')
    form.l_measurement.data = line_class_data.get('l_measurement', '')
    form.l_unit_price.data = line_class_data.get('l_unit_price', '')
    
    # Debug session data
    # print("Session before processing:", line_class_data)
    # Initialize the total variable
    total_sales = sum(
        float(item['l_sales'].replace(',', '')) for item in line_items
    ) if line_items else 0
    formatted_total_sales = f"{total_sales:,.2f}"  # Format the total for display
    print("Total sales:", formatted_total_sales)
    
    # Store total_sales in the session
    session['total_sales'] = total_sales
    # print(total_sales)
    
    if request.method == 'POST':
        # Retrieve the submitted values
        inv_no = request.form.get('l_inv_no')
        date_time = request.form.get('l_date_time')
        
        # Debug the received POST data
        print(f"Received POST data: {inv_no} {date_time}")

        # Only save to session if they are None
        if session.get('l_inv_no') is None and inv_no:
            session['l_inv_no'] = inv_no
        if session.get('l_date_time') is None and date_time:
            session['l_date_time'] = date_time

        # Debug session state after update
        print(f"Session before POST for second line item: {session.get('l_inv_no')} {session.get('l_date_time')}")

        # Ensure session is marked as modified
        session.modified = True
        print(f"Session modified: {session.modified}")

        # print("Session data updated:", session.get('l_inv_no'), session.get('l_date_time'))
        # print("Session data after POST for second line item:", session.get('l_inv_no'), session.get('l_date_time'))
        # print("Form data submitted2:", request.form.to_dict())
        
        # Get and validate quantity
        quantity_str = request.form.get('l_quantity', '0')
        
        try:
            # Attempt to convert the quantity to an integer
            quantity = int(quantity_str)
            if quantity <= 0:
            
                flash("😥 Quantity must be a positive number greater than 0.", "danger")
                # Update session while preserving existing data
                session['line_class_data'].update({
                    'l_class_code': request.form.get('l_class_code') or line_class_data.get('l_class_code'),
                    'l_desc': request.form.get('l_desc') or line_class_data.get('l_desc'),
                    'l_measurement': request.form.get('l_measurement') or line_class_data.get('l_measurement'),
                    'l_unit_price': request.form.get('l_unit_price') or line_class_data.get('l_unit_price'),
                })
                session.modified = True
                print("Session updated after invalid input:", session['line_class_data'])
                return redirect(url_for('add_line'))

        except ValueError:
            flash("😥 Quantity must be a valid number.", "danger")
            # Update session while preserving existing data
            session['line_class_data'].update({
                'l_class_code': request.form.get('l_class_code') or line_class_data.get('l_class_code'),
                'l_desc': request.form.get('l_desc') or line_class_data.get('l_desc'),
                'l_measurement': request.form.get('l_measurement') or line_class_data.get('l_measurement'),
                'l_unit_price': request.form.get('l_unit_price') or line_class_data.get('l_unit_price'),
            })
            session.modified = True
            # print("Session updated after invalid input:", session['line_class_data'])
            return redirect(url_for('add_line'))
        
        # Get unit price as string & Default to 0 if not present
        unit_price_raw= line_class_data.get('l_unit_price', '0') 

        try:
             # Handle case where unit_price is a string
            if isinstance(unit_price_raw, str): 
                 # Remove commas and convert to float
                unit_price = float(unit_price_raw.replace(',', '')) 
            else:  # Assume it's already numeric
                unit_price = float(unit_price_raw)
        except ValueError:
            flash("😥 Unit price must be a valid number.", "danger")
            return redirect(url_for('add_line'))
            
            # print(unit_price)
            
        # Perform multiplication
        sales = quantity * unit_price  
        # print(sales)
        
        # Format the total for display with commas and 2 decimal places
        formatted_total_sales = f"{sales:,.2f}"
        
        # Generate a unique ID
        new_id = len(line_items) + 1 
                
        line_data = {
            'id': str(new_id),
            'l_class_code': str(line_class_data.get('l_class_code',)),
            'l_desc': str(line_class_data.get('l_desc')),
            'l_quantity' : str(quantity),
            'l_measurement': str(line_class_data.get('l_measurement')),
            'l_unit_price': f"{unit_price:,.2f}",
            'l_sales': str(formatted_total_sales),
            }
        # print(line_data)
        # Append the new line data to the list
        line_items.append(line_data)
        # Save updated line items back to the session
        session['line_items'] = line_items
        session.modified = True  # Ensure session is updated
        
        # # Clear the session data for form fields
        session['line_class_data'] = {}  
        # # Set search_successful to False
        session['search_successful'] = False  
        flash("😄 Line item added successfully!", "success")
        
        # Redirect to the same route to clear form fields
        return redirect(url_for('add_line'))

    # Example logic to determine if the button should remain active
    search_successful = 'line_class_data' in session and session['line_class_data']
    
    # Retrieve session data
    l_inv_no = session.get('l_inv_no', '')  # Default to empty string if not set
    l_date_time = session.get('l_date_time', '')
    print(f"Retrieved session data: {l_inv_no} {l_date_time}")
    form.l_inv_no.data = l_inv_no
    form.l_date_time.data = l_date_time
    
    form.l_inv_no.render_kw = {'disabled': True}
    form.l_date_time.render_kw = {'disabled': True}
    
    # Pass the line_items list to the template
    return render_template(
    'line_items.html',
            form=form,
            line_items=line_items,
            total_sales=formatted_total_sales,
            l_inv_no=l_inv_no,
            l_date_time=l_date_time,
            search_successful=search_successful
        )

@app.route('/line_items')
def line_items():
    form = LineItemsForm()
    
    # Retrieve the line_items from session
    line_items = session.get('line_items', [])
    
    # Retrieve l_inv_no from session
    l_inv_no = session.get('l_inv_no', None)  # Default to 'INV001' if not set
    l_date_time = session.get('l_date_time', '')
    # print(f"Retrieved session data: {l_inv_no} {l_date_time}")
    
    # Populate the form fields
    form.l_inv_no.data = l_inv_no
    form.l_date_time.data = l_date_time

    # Manage field states
    if not line_items:
        form.l_inv_no.render_kw = {'disabled': False, 'style': 'background-color: #ffca2c;'}
        form.l_date_time.render_kw = {'disabled': False, 'style': 'background-color: #ffca2c;'}
        
        l_inv_no = ''
        l_date_time = ''
    else:
        form.l_inv_no.render_kw = {'disabled': True}
        form.l_date_time.render_kw = {'disabled': True}

    return render_template('line_items.html', 
                           l_inv_no=l_inv_no,
                           l_date_time=l_date_time,
                           form=form,
                           line_items=line_items)
    
@app.route('/line_clear', methods=['POST', 'GET'])
@login_required
def line_clear():
    form = LineItemsForm()
    form.l_toggle.data = ''
    form.l_class_code.data = ''
    form.l_desc.data = ''
    form.l_date_time.data = ''
    print("Clicked line clear button")
    
    # Clear form data on logout
    session.pop('line_class_data', None)  
    # Render with cleared form data instead of redirecting
    return render_template("line_items.html", 
                           form=form)
    
@app.route('/edit_line/<int:id>', methods=['GET', 'POST'])
def new_edit_line(id):
    # Initialize the form
    form = LineItemsForm()
    
    # Retrieve line items from session
    line_items = session.get('line_items', [])
    print("Session Line Items:", line_items)

    # Find the line item by ID
    line_to_edit = next((item for item in line_items if item['id'] == str(id)), None)
    print("Editing Line ID:", id)

    if not line_to_edit:
        flash("😥 Line item not found.", "danger")
        return redirect(url_for('add_line'))

    if request.method == 'GET':
        # Prepopulate the form with the selected line item data
        form.l_class_code.data = line_to_edit['l_class_code']
        form.l_desc.data = line_to_edit['l_desc']
        form.l_measurement.data = line_to_edit['l_measurement']
        form.l_unit_price.data = line_to_edit['l_unit_price']
        form.l_quantity.data = line_to_edit['l_quantity']
        
        # Disable all fields after populating them
        form.l_class_code.render_kw = {'disabled': True}
        form.l_desc.render_kw = {'disabled': True}
        form.l_measurement.render_kw = {'disabled': True}
        form.l_unit_price.render_kw = {'disabled': True}
        
    # Retrieve the same data from the session
    l_inv_no = session.get('l_inv_no', '')  # Default to empty string if not set
    l_date_time = session.get('l_date_time', '')
    print(l_inv_no, l_date_time)
    
    # Initialize the form and populate it with session data
    form.l_inv_no.data = l_inv_no
    form.l_date_time.data = l_date_time
    
    flash("😄 Please update the orange boxes.", "success")
        
    # Render the form in the template
    return render_template('edit_line.html', 
                           form=form, 
                           id=id)

@app.route('/update_line/<int:id>', methods=['POST'])
def update_line(id):
    # Initialize the form
    form = LineItemsForm()
    
    # Retrieve line items from session
    line_items = session.get('line_items', [])

    # Find the line item by ID
    line_to_update = next((item for item in line_items if item['id'] == str(id)), None)
   
    print("Update Line ID:", id)
  
    if not line_to_update:
        flash("😥 Line item not found.", "danger")
        return redirect(url_for('add_line'))

    if request.method == 'POST':
        # print("Form data submitted:", request.form.to_dict())
        
        # Retrieve the submitted values
        inv_no = request.form.get('l_inv_no')
        date_time = request.form.get('l_date_time')
        
        # Update the session
        session['l_inv_no'] = inv_no
        session['l_date_time'] = date_time

        # Get and validate quantity from the form data
        quantity_str = form.l_quantity.data
        
        # Check if the field is empty
        if not quantity_str:  
            flash("😥 Quantity cannot be left blank.", "danger")
            return redirect(url_for('edit_line', id=id))
        
        try:
            # Attempt to convert the quantity to an integer
            quantity = int(quantity_str)
            if quantity <= 0:
                flash("😥 Quantity must be a positive number greater than 0.", "danger")
                # Update session while preserving existing data
                
                # print("Session updated after invalid input:", session['line_class_data'])
                return redirect(url_for('edit_line', id=id))

        except ValueError:
            flash("😥 Quantity must be a valid number.", "danger")
            # print("Session updated after invalid input:", session['line_class_data'])
            return redirect(url_for('edit_line', id=id))
        
        # Retrieve the unit price from the session (existing value)
        unit_price_raw = line_to_update.get('l_unit_price', '0')
        
        try:
            # Handle case where unit_price is a string
            if isinstance(unit_price_raw, str): 
                # Remove commas and convert to float
                unit_price = float(unit_price_raw.replace(',', ''))  
            else:  # Assume it's already numeric
                unit_price = float(unit_price_raw)
        except ValueError:
            flash("😥 Unit price must be a valid number.", "danger")
            return redirect(url_for('edit_line', id=id))
    
        print(f"Unit price: {unit_price}")
        
        # Perform multiplication
        sales = quantity * unit_price  
        print(f"Sales: {sales}")
        
        # Format the total for display with commas and 2 decimal places
        formatted_sales = f"{sales:,.2f}"
        
        # Update the line item
        line_to_update.update({
            'l_class_code': line_to_update.get('l_class_code'),
            'l_desc': line_to_update.get('l_desc'),
            'l_measurement':line_to_update.get('l_measurement'),
            'l_unit_price': f"{unit_price:,.2f}",
            'l_quantity': quantity,
            'l_sales': formatted_sales,
        })

        # Save the updated line items list back to the session
        session['line_items'] = line_items
        session.modified = True

        flash("😄 Line item updated successfully!", "success")
        return redirect(url_for('add_line')) 
    
    # Create a form instance and populate it with session data
    form.l_inv_no.data = l_inv_no
    form.l_date_time.data = l_date_time
    
    return redirect(url_for('edit_line', 
                            id=id))

@app.route('/delete_line/<int:id>', methods=['GET','POST'])
def delete_line(id):
    # Initialize the form
    form = LineItemsForm()
    
    # Retrieve line items from the session
    line_items = session.get('line_items', [])
    
    # Filter out the item with the matching ID
    line_items = [item for item in line_items if int(item['id']) != id]
    
    # Update the session
    session['line_items'] = line_items
    session.modified = True
    
    # Check if inv_no and date_time exist in the session
    l_inv_no = session.get('l_inv_no', form.l_inv_no.data) 
    l_date_time = session.get('l_date_time', form.l_date_time.data)
    
    # Ensure these values are retained in the session
    session['l_inv_no'] = l_inv_no 
    session['l_date_time'] = l_date_time
    # print(f"Retained inv_no={l_inv_no}, date_time={l_date_time}")
    
    # Check if all line_items have been deleted 
    if not line_items:
        # Set the form fields to be enabled 
        form.l_inv_no.render_kw = {'disabled': False} 
        form.l_date_time.render_kw = {'disabled': False}
    else:
        # Disable the fields
        form.l_inv_no.render_kw = {'disabled': True} 
        form.l_date_time.render_kw = {'disabled': True}
    
    # Retain `l_inv_no` and `l_date_time` in the form
    form.l_inv_no.data = l_inv_no
    form.l_date_time.data = l_date_time
    
    flash(f"😄Item with No {id} deleted successfully!", "success")  # Flash message to notify the user

    # Redirect to the line_items page (or wherever you want to show the updated list)
    return redirect(url_for('line_items'))
    

@app.route('/save_to_excel', methods=['POST', 'GET'])
@login_required
def save_to_excel():
    # Initialize the form
    form = LineItemsForm()
    
    # Initialize the form
    # Retrieve stored session data
    inv_no = session.get('l_inv_no')
    date_time = session.get('l_date_time')
    datetime_format = datetime.strptime(date_time, "%Y-%m-%d %H:%M:%S")
    # print(f"Retrieved from session: inv_no={inv_no}, date_time={datetime_format}")
    
    ######## Supplier ######## 
    # Retrieve the line_items list from session or database
    line_items = session.get('line_items', [])
    
    s_tin = session.get('s_tin', None)
    s_name = session.get('s_name', None)
    s_id_type = session.get('s_id_type', None)
    s_business_no = session.get('s_business_no', None)
    s_email = session.get('s_email', None)
    s_msic_code = session.get('s_msic_code', None)
    s_msic_desc = session.get('s_msic_desc', None)
    s_add_1 = session.get('s_add_1', None)
    s_add_2 = session.get('s_add_2', None)
    s_add_3 = session.get('s_add_3', None)
    s_postcode = session.get('s_postcode', None)
    s_city = session.get('s_city', None)
    s_state = session.get('s_state', None)
    
    # Map states to their corresponding formatted values
    state_mapping = {
        'Johor': '01',
        'Kedah': '02',
        'Kelantan': '03',
        'Melaka': '04',
        'Negeri Sembilan': '05',
        'Pahang': '06',
        'Pulau Pinang': '07',
        'Perak': '08',
        'Perlis': '09',
        'Selangor': '10',
        'Terengganu': '11',
        'Sabah': '12',
        'Sarawak': '13',
        'Wilayah Persekutuan Kuala Lumpur': '14',
        'Wilayah Persekutuan Labuan': '15',
        'Wilayah Persekutuan Putrajaya': '16',
        'Not Applicable': '17'
    }
    formatted_state = state_mapping.get(s_state)
    
    s_country = session.get('s_country', None)
    country_mapping = {
        'Malaysia' : 'MYS',
        'Singapore':'SGP',
        'Indonesia': 'IDN',
        'Thailand': 'THA'
    }
    formatted_country = country_mapping.get(s_country)
    
    s_telephone = session.get('s_telephone', None)
    
    # print(f"'{s_add_1}', '{s_add_2}', '{s_add_3}', '{s_postcode}',\
    #       '{s_city}','{formatted_state}','{formatted_country}','{s_telephone}")
    
    ##### Process and save to Excel #####
    file_path = 'static/batch_submission/Multiple_transaction.xlsx'
    workbook = load_workbook(filename=file_path)
    print(f"'{file_path}'read!")
    
    doc_sheet = workbook["Documents"]
    start_row = 6
    # If the sheet already has data, start from the next row after the last filled row
    last_filled_row = doc_sheet.max_row
    print(last_filled_row)
    row = start_row if last_filled_row < start_row else last_filled_row + 1
    
    doc_sheet[f'B{row}'] = f'{inv_no}'
    doc_sheet[f'C{row}'] = '01'
    doc_sheet[f'D{row}'] = '1.1'
    doc_sheet[f'E{row}'] = datetime_format
    doc_sheet[f'F{row}'] = 'MYR'
    
    # 1a) Documents - supplier
    doc_sheet[f'H{row}'] = f'{s_tin}'
    doc_sheet[f'I{row}'] = f'{s_name}'
    doc_sheet[f'J{row}'] = f'{s_id_type}'
    doc_sheet[f'K{row}'] = f'{s_business_no}'
    doc_sheet[f'N{row}'] = f'{s_email}'
    doc_sheet[f'O{row}'] = s_msic_code
    doc_sheet[f'P{row}'] = f'{s_msic_desc}'
    doc_sheet[f'Q{row}'] = f'{s_add_1}'
    doc_sheet[f'R{row}'] = f'{s_add_2}'
    doc_sheet[f'S{row}'] = f'{s_add_3}'
    doc_sheet[f'T{row}'] = s_postcode
    doc_sheet[f'U{row}'] = f'{s_city}'
    doc_sheet[f'V{row}'] = f'{formatted_state}'
    doc_sheet[f'W{row}'] = f'{formatted_country}'
    doc_sheet[f'X{row}'] = f'{s_telephone}'
    workbook.save(file_path)
    
    ######## Buyer ######## 
    # Retrieve the form data dictionary from the session
    form_data = session.get('form_data', {})
    
    # Extract b_tin from the form_data dictionary
    b_tin = form_data.get('b_tin', None)
    b_name = form_data.get('b_name', None)
    b_id_type = form_data.get('b_id_type', None)
    b_id_no = form_data.get('b_id_no', None)
    b_email = form_data.get('b_email', None)
    b_add_1 = form_data.get('b_add_1', None)
    b_add_2 = form_data.get('b_add_2', None)
    b_add_3 = form_data.get('b_add_3', None)
    b_postcode = form_data.get('b_postcode', None)
    b_city = form_data.get('b_city', None)
    b_state = form_data.get('b_state', None)
    
    # Map states to their corresponding formatted values
    state_mapping1 = {
        'Johor': '01',
        'Kedah': '02',
        'Kelantan': '03',
        'Melaka': '04',
        'Negeri Sembilan': '05',
        'Pahang': '06',
        'Pulau Pinang': '07',
        'Perak': '08',
        'Perlis': '09',
        'Selangor': '10',
        'Terengganu': '11',
        'Sabah': '12',
        'Sarawak': '13',
        'Wilayah Persekutuan Kuala Lumpur': '14',
        'Wilayah Persekutuan Labuan': '15',
        'Wilayah Persekutuan Putrajaya': '16',
        'Not Applicable': '17'
    }
    formatted_state1 = state_mapping1.get(b_state)
    
    b_country = form_data.get('b_country', None)
    country_mapping1 = {
        'Malaysia' : 'MYS',
        'Singapore':'SGP',
        'Indonesia': 'IDN',
        'Thailand': 'THA'
    }
    formatted_country1 = country_mapping1.get(b_country)
    
    b_tel_no = form_data.get('b_tel_no', None)
    
    # print(f"'{formatted_state1}', '{formatted_country1}', '{b_tel_no}'")
    
    ########## 1b) Documents - buyer ##########
    doc_sheet = workbook["Documents"]
    doc_sheet[f'Y{row}'] = f'{b_tin}'
    doc_sheet[f'Z{row}'] = f'{b_name}'
    doc_sheet[f'AA{row}'] = f'{b_id_type}'
    doc_sheet[f'AB{row}'] = f'{b_id_no}'
    doc_sheet[f'AC{row}'] = 'NA'
    doc_sheet[f'AD{row}'] = f'{b_email}'
    doc_sheet[f'AE{row}'] = f'{b_add_1}'
    doc_sheet[f'AF{row}'] = f'{b_add_2}'
    doc_sheet[f'AG{row}'] = f'{b_add_3}'
    doc_sheet[f'AH{row}'] = b_postcode
    doc_sheet[f'AI{row}'] = f'{b_city}'
    doc_sheet[f'AJ{row}'] = f'{formatted_state1}'
    doc_sheet[f'AK{row}'] = f'{formatted_country1}'
    doc_sheet[f'AL{row}'] = f'{b_tel_no}'
    workbook.save(file_path)

    # Retrieve formatted_sales from the session
    total_sales = session.get('total_sales', 0)
    
    # 1b) Documents - amount
    doc_sheet = workbook["Documents"]
    doc_sheet[f'AM{row}'] = total_sales
    doc_sheet[f'AN{row}'] = total_sales
    doc_sheet[f'AO{row}'] = total_sales
    doc_sheet[f'AP{row}'] = total_sales
    doc_sheet[f'AT{row}'] = 0.00000
    
    # 5) Document Total Tax
    doc_total_tax_sheet = workbook["DocumentTotalTax"]
    last_filled_row = doc_total_tax_sheet.max_row
    total_tax_next_row = start_row if last_filled_row < start_row else last_filled_row + 1
    doc_total_tax_sheet[f'B{total_tax_next_row}'] = f'{inv_no}'
    doc_total_tax_sheet[f'C{total_tax_next_row}'] = "06"
    doc_total_tax_sheet[f'E{total_tax_next_row}'] = 0
    
    # 7) Document Discounts 
    doc_dis_sheet = workbook["DocumentDiscounts"]
    last_filled_row = doc_dis_sheet.max_row
    doc_dis_next_row = start_row if last_filled_row < start_row else last_filled_row + 1
    doc_dis_sheet[f'B{doc_dis_next_row}'] = f'{inv_no}'
    doc_dis_sheet[f'C{doc_dis_next_row}'] = 'no'
    doc_dis_sheet[f'D{doc_dis_next_row}'] = 0
    
    # 9) Document Charges
    doc_fee_sheet = workbook["DocumentCharges"]
    last_filled_row = doc_fee_sheet.max_row
    doc_fee_next_row = start_row if last_filled_row < start_row else last_filled_row + 1
    doc_fee_sheet[f'B{doc_fee_next_row}'] = f'{inv_no}'
    doc_fee_sheet[f'C{doc_fee_next_row}'] = 'no'
    doc_fee_sheet[f'D{doc_fee_next_row}'] = 0
    workbook.save(file_path)
    
    ########## 2) Document Line Items ##########
   # Retrieve line_items from the session, default to an empty list if not present
    line_items = session.get('line_items', [])
    # print(line_items)
    
    start_row = 6
    max_rows = 100
    
    doc_line_sheet = workbook["DocumentLineItems"]
    # If the sheet already has data, start from the next row after the last filled row
    last_filled_row = doc_line_sheet.max_row
    print('last row', f'{last_filled_row}')
    next_row = start_row if last_filled_row < start_row else last_filled_row + 1
    print('next row', f'{next_row}')
    
    for idx, item in enumerate(line_items[:max_rows]):
        ########## Read from session ##########
        l_id = item['id']
        l_class_code = item['l_class_code']
        l_class_code_formatted = l_class_code.split('-')[0].strip()
        l_desc = item['l_desc']
        l_unit_price = item['l_unit_price']
        l_measurement = item['l_measurement']
        
        measurement_mapping = {
                'Box' : 'XBX',
                'day':'DAY',
                'job': 'E51',
                'month': 'MON',
                'Package':'XPK',
                'piece': 'XPP',
                'set': 'set',
                'trip':'E54',
                'week': 'WEE',
                'Unit': 'XUN',
                'each': 'EA'
            }
        formatted_measurement = measurement_mapping.get(l_measurement)
        print(formatted_measurement)
        
        l_quantity = item['l_quantity']
        l_sales = item['l_sales']
        
        # Debug print to confirm the data being written
        # pprint(f"Row {row} - ID: {l_id}, Class Code: {l_class_code_formatted}, Desc: {l_desc}, \
        #        Quantity: {l_quantity}, Price: {l_unit_price}, Total Sales: {l_sales}")

        ########## Write to Excel Sheet ##########
        # 2) Document Line Items
        doc_line_sheet = workbook["DocumentLineItems"]
        last_filled_row = doc_line_sheet.max_row
        doc_next_row = start_row if last_filled_row < start_row else last_filled_row + 1
        doc_line_sheet[f'B{doc_next_row}'] = f'{inv_no}'
        doc_line_sheet[f'C{doc_next_row}'] = l_id
        doc_line_sheet[f'D{doc_next_row}'] = l_class_code_formatted
        doc_line_sheet[f'E{doc_next_row}'] = l_desc
        doc_line_sheet[f'F{doc_next_row}'] = l_unit_price
        doc_line_sheet[f'G{doc_next_row}'] = l_quantity
        doc_line_sheet[f'H{doc_next_row}'] = formatted_measurement
        doc_line_sheet[f'I{doc_next_row}'] = l_sales
        doc_line_sheet[f'J{doc_next_row}'] = l_sales
        doc_line_sheet[f'K{doc_next_row}'] = l_sales
        
        # 3) Line Items Add Classifications
        line_items_class_sheet = workbook["LineItemsAddClassifications"]
        last_filled_row = line_items_class_sheet.max_row
        class_next_row = start_row if last_filled_row < start_row else last_filled_row + 1
        line_items_class_sheet[f'B{class_next_row}'] = f'{inv_no}'
        line_items_class_sheet[f'C{class_next_row}'] = l_id
        line_items_class_sheet[f'D{class_next_row}'] = l_class_code_formatted
        
        # 4) Line Items Taxes
        line_items_tax_sheet = workbook["LineItemsTaxes"]
        last_filled_row = line_items_tax_sheet.max_row
        tax_next_row = start_row if last_filled_row < start_row else last_filled_row + 1
        line_items_tax_sheet[f'B{tax_next_row}'] = f'{inv_no}'
        line_items_tax_sheet[f'C{tax_next_row}'] = l_id
        line_items_tax_sheet[f'D{tax_next_row}'] = "06"
        line_items_tax_sheet[f'F{tax_next_row}'] = 0
        
        # 6) Line Items Discounts
        line_item_dis_sheet = workbook["LineItemsDiscounts"]
        last_filled_row = line_item_dis_sheet.max_row
        dis_next_row = start_row if last_filled_row < start_row else last_filled_row + 1
        line_item_dis_sheet[f'C{dis_next_row}'] = l_id
        line_item_dis_sheet[f'B{dis_next_row}'] = f'{inv_no}'
        line_item_dis_sheet[f'D{dis_next_row}'] = 0.0
        line_item_dis_sheet[f'E{dis_next_row}'] =  0
        line_item_dis_sheet[f'F{dis_next_row}'] = 'no'
        
        # 8) Line Items Charges
        line_item_fee_sheet = workbook["LineItemsCharges"]
        last_filled_row = line_item_fee_sheet.max_row
        fee_next_row = start_row if last_filled_row < start_row else last_filled_row + 1
        line_item_fee_sheet[f'B{fee_next_row}'] = f'{inv_no}'
        line_item_fee_sheet[f'C{fee_next_row}'] = l_id
        line_item_fee_sheet[f'D{fee_next_row}'] = 0.0
        line_item_fee_sheet[f'E{fee_next_row}'] =  0
        line_item_fee_sheet[f'F{fee_next_row}'] = 'No'
           
    workbook.save(file_path)
     
    # Retrieve session data
    l_inv_no = session.get('l_inv_no', '')  # Default to empty string if not set
    l_date_time = session.get('l_date_time', '')
    
    # Create a form instance and populate it with session data
    form.l_inv_no.data = l_inv_no
    form.l_date_time.data = l_date_time
    
    fields_disabled = session.get('fields_disabled', False)
    form.l_inv_no.render_kw = {'disabled': fields_disabled}
    form.l_date_time.render_kw = {'disabled': fields_disabled}
    
    # Embed HTML with buttons in the flash message
    flash(
    """
    <div>
        <p>Excel written successfully. You may either:</p> 
        <a href="/download_excel" class="btn btn-primary">Download Excel</a>
        <a href="/start_new_buyer" class="btn btn-secondary">Start A New Buyer</a>
        <p><small>Before batch upload to MyInvois Portal, please don't forget to save the excel file!</small></p>
    </div>
    """,
    "success")
    
    return render_template('line_items.html', 
                           form=form, 
                           line_items=line_items,)
        
@app.route('/line_search', methods=['POST', 'GET'])
@login_required
def line_search():
    # Initialize search_successful to False
    # search_successful = False 
    session['search_successful'] = False
    # Ensure it's initialized
    search_successful = False 
    # print("line_search function called")  # Debugging
    form = LineItemsForm(is_search=True)
    
    ########## Inv_no & date_time #########
    # Check if values exist in the session
    l_inv_no = session.get('l_inv_no', '')  # Default to empty string if not set
    l_date_time = session.get('l_date_time', '')
    
    # Populate the form fields
    form.l_inv_no.data = l_inv_no
    form.l_date_time.data = l_date_time
    print(f'Line Search: {l_inv_no},{l_date_time}')
    
    # Adjust field attributes based on session values
    if l_inv_no and l_date_time:
        # If session values are set, disable the fields
        form.l_inv_no.render_kw = {'disabled': True} 
    else:
        # If session values are not set, provide default values and a custom background color
        form.l_inv_no.render_kw = {'style': 'background-color: #ffca2c;'}
    
    ##########################
    # Default to no results found
    line_result = None

    if form.validate_on_submit():
        # Extract the search term from the form field
        line_search = form.l_search.data 
        # print(f"Searching for: {line_search}")  # Debugging
        
        line_result = form.l_result.data
        # print(f"Searching for: {line_result}")
        
        if not line_search:
            flash("😥 Please fill in the search input", "danger")
            return redirect(url_for('line_items'))
        
        # Load the Excel file
        excel_files = glob('static/upload/*.xlsx')
        
        if not excel_files:
            flash("😥 No Excel files found", "danger")
            return redirect(url_for('line_items'))
    
        # Use the first Excel file found in the folder
        file_path = excel_files[0]
        
        try:
            # Try to load the Excel file
            df = pd.read_excel(file_path, sheet_name=2)
            # Filter the DataFrame by the search term in either "Alias" or "Name*" columns
            # result = df[df['Alias'].astype(str).str.contains(str(line_search), case=False, na=False) |
            #          df['Description of Product or Service*'].astype(str).str.contains(str(line_search), case=False, na=False)]
            df['Alias'] = df['Alias'].astype(str)
            df['Description of Product or Service*'] = df['Description of Product or Service*'].apply(
                lambda x: str(int(x)) if isinstance(x, (int, float)) and not pd.isna(x) else str(x)
            )

            line_search = str(int(line_search)) if line_search.isdigit() else str(line_search)

            result = df[df['Alias'].str.contains(line_search, case=False, na=False) |
                        df['Description of Product or Service*'].str.contains(line_search, case=False, na=False)]

            # print(result)
            
             # Check for duplicate names in the results
            if len(result) > 1:
                flash("😥 Duplicate name found. \
                      Please input the correct alias of description of product and service!", "danger")
                return redirect(url_for('line_items'))
            
            # Check if there are any results 
            if not result.empty:
                # Accessing the first match
                class_code = result['Classification Codes* (Drop box)'].iloc[0]  
                # print(f"Classification Codes No.: {class_code}")
                
                prod_serv = result['Description of Product or Service*'].iloc[0] 
                # print(f"Description of Product or Service*: {prod_serv}")
                
                measurement = result['Measurement* (Drop box)'].iloc[0]
                # print(f"Measurement*: {measurement}")
                
                unit_price = int(result['Unit Price*(RM)'].iloc[0]) 
                # print(f"Unit Price*(RM): {unit_price}")
                
                # Set each form field's data
                form.l_class_code.data = class_code
                form.l_desc.data = prod_serv
                form.l_measurement.data = measurement
                form.l_unit_price.data = unit_price
                form.l_result.data = line_result  # Set form data
                
                # Disable all fields after populating them
                form.l_class_code.render_kw = {'disabled': True}
                form.l_desc.render_kw = {'disabled': True}
                form.l_measurement.render_kw = {'disabled': True}
                form.l_unit_price.render_kw = {'disabled': True}
                
                # Store data in session
                session['line_class_data'] = {
                    'l_class_code': class_code,
                    'l_desc': prod_serv,
                    'l_measurement': measurement,
                    'l_unit_price': unit_price,
                    'l_search': line_search,  # store search term
                }
                # print(session['line_class_data'])
              
                # Extract the name for the supplier and buyer display
                df = pd.read_excel(file_path, sheet_name=0)
                s_name = df.iloc[0, 1]
                
                df = pd.read_excel(file_path, sheet_name=1)
                b_name = df.iloc[0, 4]
                
                # Set `l_result` with both names
                form.l_result.data = f"Supplier name:  {s_name}\nBuyer name:  {b_name}\nLine Item: {prod_serv}"
                
                # Store `search_successful` in session
                session['search_successful'] = True 
                
                # Set the success flag to True to enable the Next button
                search_successful = True
                # print(f"Search successful, result: {form.l_result.data}")  # Debugging
            
                flash("😄 Alright your search found. Please fill out the orange boxes and\
                      click the 'Add Line' button!", "success")
            else:         
           
                form.l_result.data = "No matches found."
                # Clear session if no results found
                session.pop('search_successful', None)  
                # Reset the flag
                search_successful = False  
    
        except Exception as e:
            flash(f"😥 Error reading Excel file: {e}", "danger")
        
    return render_template('line_items.html', 
                            form=form, 
                            l_inv_no=l_inv_no,
                            l_date_time=l_date_time,
                            search_successful=search_successful,
                            line_result=line_result)
    
@app.route('/additional_info', methods=['GET', 'POST'])
def additional_info():
    form = AddInfoForm()
    if request.method == 'POST' and form.validate():
        discount_description = (
            form.a_dis_desc_custom.data
            if form.a_dis_desc.data == 'other'
            else form.a_dis_desc.data
        )
        fee_description = (
            form.a_fee_desc_custom.data
            if form.a_fee_desc.data == 'other'
            else form.a_fee_desc.data
        )
        
        flash(f"😄 Discount Description: {discount_description}")
        flash(f"😄 Fee/Charge Description: {fee_description}")
        
        try:
            discount_value = form.a_dis_value.data  # This will be a Decimal
            # Process the discount_value as needed
            return f"Discount value processed: {discount_value}"
        except Exception as e:
            return f"Error: {e}"
        
        prepayment_num = form.a_prepayment_num.data
        return f"Prepayment Reference Number: {prepayment_num}"
    
    return render_template('additional_info.html', 
                           form=form)

@app.route('/additional_save_to_excel', methods=['GET', 'POST'])
def additional_save_to_excel():
    # Initialize the form
    form = AddInfoForm()
  
    if request.method == 'POST':
        # Print the submitted data to debug
        # print(request.form)

        # # Retrieve values with proper fallback
        dis_desc = request.form.get('a_dis_desc', '').strip()
        dis_desc_custom = request.form.get('a_dis_desc_custom', '').strip()
        
        # # Check if the selected description is 'Other', and use the custom description
        if dis_desc == 'other' and dis_desc_custom:
            dis_desc = dis_desc_custom
        elif dis_desc != 'other':
            # Otherwise use the selected description
            dis_desc = dis_desc
        # print(dis_desc)
        
        dis_value = request.form.get('a_dis_value', '').strip()
        
        fee_desc = request.form.get('a_fee_desc', '').strip()
        fee_desc_custom = request.form.get('a_fee_desc_custom', '').strip()
        
        #  # Check if the selected description is 'Other', and use the custom description
        if fee_desc == 'other' and fee_desc_custom:
            fee_desc = fee_desc_custom
        elif fee_desc != 'other':
            # Otherwise use the selected description
            fee_desc = fee_desc
        # print(fee_desc)
        
        fee_value = request.form.get('a_fee_value', '').strip()
        
        payment_mode = request.form.get('a_payment_mode', '').strip()
        
        # # Map states to their corresponding formatted values
        payment_code_mapping = {
            'NA':'',
            'Cash': '01',
            'Cheque': '02',
            'Bank Transfer': '03',
            'Credit Card': '04',
            'Debit Card': '05',
            'E-wallet / Digital Wallet': '06',
            'Digital Bank': '07',
            'Others': '08',
        }
        formatted_payment_code = payment_code_mapping.get(payment_mode)
        print(formatted_payment_code)
        
        bank_acc = request.form.get('a_bank_acc', '').strip()
        payment_term = request.form.get('a_payment_term', '').strip()
        payment_amount = request.form.get('a_payment_amount', '').strip()
        
        try:
            dis_value = float(dis_value) if dis_value else 0.0
            fee_value = float(fee_value) if fee_value else 0.0
            payment_amount = float(payment_amount) if payment_amount else 0.0
        except ValueError:
            flash('Please enter valid numbers for discount and fee values.', 'error')
            return render_template('additional_info.html', form=form, total_sales=total_sales)
        
        prepayment_date = request.form.get('a_prepayment_date', '').strip()
        print(prepayment_date)
        prepayment_num = request.form.get('a_prepayment_num', '').strip()
        bill_num = request.form.get('a_bill_num', '').strip()

        # Debugging output
        # print(f"Discount Description: '{dis_desc}', Custom: '{dis_desc_custom}', Value: '{dis_value}'")
        # print(f"Fee Description: {fee_desc or 'Not provided'}, Custom: {fee_desc_custom or 'Not provided'}, Value: {fee_value or 'Not provided'}")
        # print(f"Payment Mode: {payment_mode or 'Not provided'}, Bank Account: {bank_acc or 'Not provided'}")
        # print(f"Payment Term: {payment_term or 'Not provided'}, Payment Amount: {payment_amount or 'Not provided'}")
        # print(f"Prepayment Date: {prepayment_date or 'Not provided'}, Prepayment Number: {prepayment_num or 'Not provided'}, Bill Number: {bill_num or 'Not provided'}")    
            
        # Initialize the form
        # Retrieve stored session data
        inv_no = session.get('l_inv_no')
        date_time = session.get('l_date_time')
        datetime_format = datetime.strptime(date_time, "%Y-%m-%d %H:%M:%S")
        # print(f"Retrieved from session: inv_no={inv_no}, date_time={datetime_format}")
        
        ######## Supplier ######## 
        # Retrieve the line_items list from session or database
        line_items = session.get('line_items', [])
        
        s_tin = session.get('s_tin', None)
        s_name = session.get('s_name', None)
        s_id_type = session.get('s_id_type', None)
        s_business_no = session.get('s_business_no', None)
        s_email = session.get('s_email', None)
        s_msic_code = session.get('s_msic_code', None)
        s_msic_desc = session.get('s_msic_desc', None)
        s_add_1 = session.get('s_add_1', None)
        s_add_2 = session.get('s_add_2', None)
        s_add_3 = session.get('s_add_3', None)
        s_postcode = session.get('s_postcode', None)
        s_city = session.get('s_city', None)
        s_state = session.get('s_state', None)
        
        # Map states to their corresponding formatted values
        state_mapping = {
            'Johor': '01',
            'Kedah': '02',
            'Kelantan': '03',
            'Melaka': '04',
            'Negeri Sembilan': '05',
            'Pahang': '06',
            'Pulau Pinang': '07',
            'Perak': '08',
            'Perlis': '09',
            'Selangor': '10',
            'Terengganu': '11',
            'Sabah': '12',
            'Sarawak': '13',
            'Wilayah Persekutuan Kuala Lumpur': '14',
            'Wilayah Persekutuan Labuan': '15',
            'Wilayah Persekutuan Putrajaya': '16',
            'Not Applicable': '17'
        }
        formatted_state = state_mapping.get(s_state)
        
        s_country = session.get('s_country', None)
        country_mapping = {
            'Malaysia' : 'MYS',
            'Singapore':'SGP',
            'Indonesia': 'IDN',
            'Thailand': 'THA'
        }
        formatted_country = country_mapping.get(s_country)
        
        s_telephone = session.get('s_telephone', None)
        
        # print(f"'{s_add_1}', '{s_add_2}', '{s_add_3}', '{s_postcode}',\
        #       '{s_city}','{formatted_state}','{formatted_country}','{s_telephone}")
        
        ##### Process and save to Excel #####
        file_path = 'static/batch_submission/Multiple_transaction.xlsx'
        workbook = load_workbook(filename=file_path)
        # print(f"'{file_path}'read!")
        
        doc_sheet = workbook["Documents"]
        start_row = 6
        # If the sheet already has data, start from the next row after the last filled row
        last_filled_row = doc_sheet.max_row
        print(last_filled_row)
        row = start_row if last_filled_row < start_row else last_filled_row + 1
        
        doc_sheet[f'B{row}'] = f'{inv_no}'
        doc_sheet[f'C{row}'] = '01'
        doc_sheet[f'D{row}'] = '1.1'
        doc_sheet[f'E{row}'] = datetime_format
        doc_sheet[f'F{row}'] = 'MYR'
        
        # 1a) Documents - supplier
        doc_sheet[f'H{row}'] = f'{s_tin}'
        doc_sheet[f'I{row}'] = f'{s_name}'
        doc_sheet[f'J{row}'] = f'{s_id_type}'
        doc_sheet[f'K{row}'] = f'{s_business_no}'
        doc_sheet[f'N{row}'] = f'{s_email}'
        doc_sheet[f'O{row}'] = s_msic_code
        doc_sheet[f'P{row}'] = f'{s_msic_desc}'
        doc_sheet[f'Q{row}'] = f'{s_add_1}'
        doc_sheet[f'R{row}'] = f'{s_add_2}'
        doc_sheet[f'S{row}'] = f'{s_add_3}'
        doc_sheet[f'T{row}'] = s_postcode
        doc_sheet[f'U{row}'] = f'{s_city}'
        doc_sheet[f'V{row}'] = f'{formatted_state}'
        doc_sheet[f'W{row}'] = f'{formatted_country}'
        doc_sheet[f'X{row}'] = f'{s_telephone}'
        workbook.save(file_path)
        
        ######## Buyer ######## 
        # Retrieve the form data dictionary from the session
        form_data = session.get('form_data', {})
        
        # Extract b_tin from the form_data dictionary
        b_tin = form_data.get('b_tin', None)
        b_name = form_data.get('b_name', None)
        b_id_type = form_data.get('b_id_type', None)
        b_id_no = form_data.get('b_id_no', None)
        b_email = form_data.get('b_email', None)
        b_add_1 = form_data.get('b_add_1', None)
        b_add_2 = form_data.get('b_add_2', None)
        b_add_3 = form_data.get('b_add_3', None)
        b_postcode = form_data.get('b_postcode', None)
        b_city = form_data.get('b_city', None)
        b_state = form_data.get('b_state', None)
        
        # Map states to their corresponding formatted values
        state_mapping1 = {
            'Johor': '01',
            'Kedah': '02',
            'Kelantan': '03',
            'Melaka': '04',
            'Negeri Sembilan': '05',
            'Pahang': '06',
            'Pulau Pinang': '07',
            'Perak': '08',
            'Perlis': '09',
            'Selangor': '10',
            'Terengganu': '11',
            'Sabah': '12',
            'Sarawak': '13',
            'Wilayah Persekutuan Kuala Lumpur': '14',
            'Wilayah Persekutuan Labuan': '15',
            'Wilayah Persekutuan Putrajaya': '16',
            'Not Applicable': '17'
        }
        formatted_state1 = state_mapping1.get(b_state)
        
        b_country = form_data.get('b_country', None)
        country_mapping1 = {
            'Malaysia' : 'MYS',
            'Singapore':'SGP',
            'Indonesia': 'IDN',
            'Thailand': 'THA'
        }
        formatted_country1 = country_mapping1.get(b_country)
        
        b_tel_no = form_data.get('b_tel_no', None)
        
        # print(f"'{formatted_state1}', '{formatted_country1}', '{b_tel_no}'")
        
        ########## 1b) Documents - buyer ##########
        doc_sheet = workbook["Documents"]
        doc_sheet[f'Y{row}'] = f'{b_tin}'
        doc_sheet[f'Z{row}'] = f'{b_name}'
        doc_sheet[f'AA{row}'] = f'{b_id_type}'
        doc_sheet[f'AB{row}'] = f'{b_id_no}'
        doc_sheet[f'AC{row}'] = 'NA'
        doc_sheet[f'AD{row}'] = f'{b_email}'
        doc_sheet[f'AE{row}'] = f'{b_add_1}'
        doc_sheet[f'AF{row}'] = f'{b_add_2}'
        doc_sheet[f'AG{row}'] = f'{b_add_3}'
        doc_sheet[f'AH{row}'] = b_postcode
        doc_sheet[f'AI{row}'] = f'{b_city}'
        doc_sheet[f'AJ{row}'] = f'{formatted_state1}'
        doc_sheet[f'AK{row}'] = f'{formatted_country1}'
        doc_sheet[f'AL{row}'] = f'{b_tel_no}'
        workbook.save(file_path)

        # Retrieve formatted_sales from the session
        total_sales = session.get('total_sales', 0)
        # print(total_sales)
        net_sales = total_sales - dis_value + fee_value
        
        amount_payable = net_sales -  payment_amount
        # print(amount_payable)
        
        # 1b) Documents - amount
        doc_sheet = workbook["Documents"]
        doc_sheet[f'AM{row}'] = net_sales #<=== total sales - discount + charges
        doc_sheet[f'AN{row}'] = net_sales #<=== total sales - discount + charges
        doc_sheet[f'AO{row}'] = amount_payable #<=== net sales - prepayment 
        doc_sheet[f'AP{row}'] = total_sales
        doc_sheet[f'AT{row}'] = 0.00000
        doc_sheet[f'AX{row}'] = f'{formatted_payment_code}'
        doc_sheet[f'AY{row}'] = f'{bank_acc}'
        doc_sheet[f'AZ{row}'] = payment_term
        doc_sheet[f'BA{row}'] = payment_amount
        doc_sheet[f'BB{row}'] = prepayment_date
        doc_sheet[f'BD{row}'] = prepayment_num
        doc_sheet[f'BE{row}'] = bill_num
        
        # 5) Document Total Tax -no looping
        doc_total_tax_sheet = workbook["DocumentTotalTax"]
        doc_total_tax_sheet[f'B{row}'] = f'{inv_no}'
        doc_total_tax_sheet[f'C{row}'] = "06"
        doc_total_tax_sheet[f'E{row}'] = 0
        
        # 7) Document Discounts - no looping
        doc_dis_sheet = workbook["DocumentDiscounts"]
        doc_dis_sheet[f'B{row}'] = f'{inv_no}'
        doc_dis_sheet[f'C{row}'] = f'{dis_desc}'
        doc_dis_sheet[f'D{row}'] = dis_value
        
        # 9) Document Charges - no looping
        doc_fee_sheet = workbook["DocumentCharges"]
        doc_fee_sheet[f'B{row}'] = f'{inv_no}'
        doc_fee_sheet[f'C{row}'] = f'{fee_desc}'
        doc_fee_sheet[f'D{row}'] = fee_value
        workbook.save(file_path)
        
        ########## 2) Document Line Items ##########
        # Retrieve line_items from the session, default to an empty list if not present
        line_items = session.get('line_items', [])
        # print(line_items)
        
        start_row = 6
        max_rows = 100
        
        doc_line_sheet = workbook["DocumentLineItems"]
        # If the sheet already has data, start from the next row after the last filled row
        last_filled_row = doc_line_sheet.max_row
        # print('last row', f'{last_filled_row}')
        next_row = start_row if last_filled_row < start_row else last_filled_row + 1
        # print('next row', f'{next_row}')
        
        for idx, item in enumerate(line_items[:max_rows]):
            ########## Read from session ##########
            l_id = item['id']
            l_class_code = item['l_class_code']
            l_class_code_formatted = l_class_code.split('-')[0].strip()
            l_desc = item['l_desc']
            l_unit_price = item['l_unit_price']
            l_quantity = item['l_quantity']
            l_measurement = item['l_measurement']
            
            measurement_mapping = {
                'Box' : 'XBX',
                'day':'DAY',
                'job': 'E51',
                'month': 'MON',
                'Package':'XPK',
                'piece': 'XPP',
                'set': 'set',
                'trip':'E54',
                'week': 'WEE',
                'Unit': 'XUN',
                'each': 'EA'
            }
            formatted_measurement = measurement_mapping.get(l_measurement)
            # print(formatted_measurement)
            
            l_sales = item['l_sales']
            
            # Debug print to confirm the data being written
            # pprint(f"Row {row} - ID: {l_id}, Class Code: {l_class_code_formatted}, Desc: {l_desc}, \
            #        Quantity: {l_quantity}, Price: {l_unit_price}, Total Sales: {l_sales}")

            ########## Write to Excel Sheet ##########
            # 2) Document Line Items
            doc_line_sheet = workbook["DocumentLineItems"]
            last_filled_row = doc_line_sheet.max_row
            print("last filled row", last_filled_row)
            doc_next_row = start_row if last_filled_row < start_row else last_filled_row + 1
            print("next row", doc_next_row)
            doc_line_sheet[f'B{doc_next_row}'] = f'{inv_no}'
            doc_line_sheet[f'C{doc_next_row}'] = l_id
            doc_line_sheet[f'D{doc_next_row}'] = l_class_code_formatted
            doc_line_sheet[f'E{doc_next_row}'] = l_desc
            doc_line_sheet[f'F{doc_next_row}'] = l_unit_price
            doc_line_sheet[f'G{doc_next_row}'] = l_quantity
            doc_line_sheet[f'H{doc_next_row}'] = formatted_measurement
            doc_line_sheet[f'I{doc_next_row}'] = l_sales
            doc_line_sheet[f'J{doc_next_row}'] = l_sales
            doc_line_sheet[f'K{doc_next_row}'] = l_sales
            
            # 3) Line Items Add Classifications
            line_items_class_sheet = workbook["LineItemsAddClassifications"]
            last_filled_row = line_items_class_sheet.max_row
            class_next_row = start_row if last_filled_row < start_row else last_filled_row + 1
            line_items_class_sheet[f'B{class_next_row}'] = f'{inv_no}'
            line_items_class_sheet[f'C{class_next_row}'] = l_id
            line_items_class_sheet[f'D{class_next_row}'] = l_class_code_formatted
            
            # 4) Line Items Taxes
            line_items_tax_sheet = workbook["LineItemsTaxes"]
            last_filled_row = line_items_tax_sheet.max_row
            tax_next_row = start_row if last_filled_row < start_row else last_filled_row + 1
            line_items_tax_sheet[f'B{tax_next_row}'] = f'{inv_no}'
            line_items_tax_sheet[f'C{tax_next_row}'] = l_id
            line_items_tax_sheet[f'D{tax_next_row}'] = "06"
            line_items_tax_sheet[f'F{tax_next_row}'] = 0
            
            # 6) Line Items Discounts
            line_item_dis_sheet = workbook["LineItemsDiscounts"]
            last_filled_row = line_item_dis_sheet.max_row
            dis_next_row = start_row if last_filled_row < start_row else last_filled_row + 1
            line_item_dis_sheet[f'C{dis_next_row}'] = l_id
            line_item_dis_sheet[f'B{dis_next_row}'] = f'{inv_no}'
            line_item_dis_sheet[f'D{dis_next_row}'] = 0.0
            line_item_dis_sheet[f'E{dis_next_row}'] =  0
            line_item_dis_sheet[f'F{dis_next_row}'] = 'no'
               
            # 8) Line Items Charges
            line_item_fee_sheet = workbook["LineItemsCharges"]
            last_filled_row = line_item_fee_sheet.max_row
            fee_next_row = start_row if last_filled_row < start_row else last_filled_row + 1
            line_item_fee_sheet[f'B{fee_next_row}'] = f'{inv_no}'
            line_item_fee_sheet[f'C{fee_next_row}'] = l_id
            line_item_fee_sheet[f'D{fee_next_row}'] = 0.0
            line_item_fee_sheet[f'E{fee_next_row}'] =  0
            line_item_fee_sheet[f'F{fee_next_row}'] = 'No'
            
        workbook.save(file_path)
        
        # Embed HTML with buttons in the flash message
        flash(
        """
        <div>
            <p>😄 Excel written successfully. You may either:</p> 
            <a href="/download_excel" class="btn btn-primary">Download Excel</a>
            <a href="/start_new_buyer" class="btn btn-secondary">Start A New Buyer</a>
            <p><small>Before batch upload to MyInvois Portal, please don't forget to save the excel file!</small></p>
        </div>
        """,
        "success")
        
    return render_template('additional_info.html', 
                           form=form, 
                           line_items=line_items,)
    
@app.route('/admin', methods=['GET', 'POST'])
def admin():
    # Query the data
    users = Users.query.all()
    
    # Prepare headers based on your data structure
    header = ['ID', 'Username', 'Email', 'Password Hash', 'Blacklisted']
    
    content = [(user.id, user.username, user.email, user.password_hash, user.blacklisted) for user in users]
    
    # Connect to the PostgreSQL database
    engine = create_engine(os.environ.get("DATABASE_URL"))
    # Fetch data from the `users` table
    with engine.connect() as my_conn:
        result = my_conn.execute(text("SELECT * FROM users"))
        
        # Check if the query was successful and contains data
        if result is not None:
            # Extract header (column names) and content (rows)
            header = list(result.keys())
            content = [list(row) for row in result]
            
            # Check if the relevant columns exist
            if "password_hash" in header and "blacklisted" in header:
                password_index = header.index("password_hash")
                blacklist_index = header.index("blacklisted")

                # Truncate the `password_hash` column for display
                # for row in content:
                #     row[password_index] = row[password_index][:5] + "..."  # Truncate password_hash
            else:
                print("Required columns missing from the table schema")
        else:
            # If result is None, handle it gracefully
            flash("Failed to retrieve users from the database", 'danger')
            content = []  # Empty content

    return render_template('admin.html', 
                           header=header, 
                           content=content)

   
from flask import flash, redirect, url_for

@app.route('/update_blacklist', methods=['POST'])
def update_blacklist():
    # Loop through form data to handle each user's checkbox update
    for key, value in request.form.items():
        if key.startswith('blacklisted_'):  # Identifying the blacklisted checkboxes
            user_id = key.split('_')[1]  # Extract user ID from the key (e.g., blacklisted_1, blacklisted_2)
            print(f"User ID: {user_id}, Blacklisted: {value}")
            blacklisted = bool(value)  # If checked, value will be '1'; otherwise '0'
            
            # Query the user by ID
            user = Users.query.get(user_id)
            if user:
                user.blacklisted = blacklisted  # Update blacklisted status
                db.session.commit()
                flash(f"User {user.username} has been {'blacklisted' if blacklisted else 'unblacklisted'} successfully!", "success")
            else:
                flash(f"User with ID {user_id} not found.", "danger")
                
    return redirect(url_for('admin'))  # Redirect to admin page after updating

@app.route('/delete_user/<int:id>', methods=['GET'])
def delete_user(id):
    # Query the user by ID
    user = Users.query.get(id)
    
    if user:
        # Delete the user from the database
        db.session.delete(user)
        db.session.commit()
        flash('User successfully deleted!', 'success')
    else:
        flash('User not found!', 'danger')

    # Redirect to the page where the users are listed (e.g., users dashboard)
    return redirect(url_for('admin'))

@app.route('/users')
def users():
    # Query all users
    users_list = Users.query.all()

    # Pass the users list to the template
    return render_template('admin.html', data=users_list)

@app.route('/robots.txt')
def robots():
    r = Response(response="User-Agent: *\nDisallow: \nSitemap: https://how-e-invoice.onrender.com/sitemap.xml", 
                 status=200, mimetype="text/plain")
    r.headers["Content-Type"] = "text/plain; charset=utf-8"
    return r

@app.route('/sitemap.xml')
def sitemap():
    urls = [
        'https://how-e-invoice.onrender.com/',
        'https://how-e-invoice.onrender.com//sign_up',
        'https://how-e-invoice.onrender.com//search',
        'https://how-e-invoice.onrender.com/line_items',
        'https://how-e-invoice.onrender.com/add_line',
        'https://how-e-invoice.onrender.com//save_to_excel',
        'https://how-e-invoice.onrender.com/additional_info',
        'https://how-e-invoice.onrender.com/additional_save_to_excel'
    ]
    
    sitemap_xml = """<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
"""
    for url in urls:
        sitemap_xml += f"""
    <url>
        <loc>{url}</loc>
        <lastmod>{datetime.utcnow().date().isoformat()}</lastmod>
        <changefreq>weekly</changefreq>
        <priority>0.8</priority>
    </url>
"""
    sitemap_xml += "</urlset>"

    # Explicitly set the Content-Type header
    response = Response(sitemap_xml)
    response.headers["Content-Type"] = "application/xml"
    
    return response

@app.route('/ads.txt')
def serve_ads_txt():
    r = Response(response="google.com, pub-4375949754721942, DIRECT, f08c47fec0942fa0",
                 mimetype="text/plain")
    return r

if __name__ == '__main__':
    #DEBUG is SET to TRUE. CHANGE FOR PROD
    app.run(port=5000,debug=True)
    

'''
pip freeze > requirements.txt   
clear
.venv\Scripts\activate

python.exe -m pip list

flask db migrate -m "Add user_agreement_timestamp to Users"

1) If the remote repository is not already set up, add it using:
git remote add origin https://github.com/Kelvin-Data/How-invoice.git

2) If origin is already set, you can skip this step or verify the URL with:
git remote -v

3) Add all files to the staging area:
git add .

4) Commit the changes with a message:
git commit -m "Your commit message here"

5) Push the changes to the remote repository:
git push -u origin main

git status

mkvirtualenv myvirtualenv --python=/usr/bin/python3.10
'''
